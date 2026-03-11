"""
build_excel.py
==============
Regenerates both Excel TEA models (Steel + Cement).
Run this if you want to rebuild the Excel files from scratch.

Usage:
    python scripts/build_excel.py
"""

import subprocess
import sys
import os

def check_openpyxl():
    try:
        import openpyxl
        print(f"  ✅ openpyxl {openpyxl.__version__} found")
    except ImportError:
        print("  ❌ openpyxl not found. Run: pip install -r requirements.txt")
        sys.exit(1)

def build_steel():
    print("\n  Building Steel TEA Excel...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Import the steel model
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'models'))
    from steel_tea import (
        total_heat_kwh, energy_input_kwh, electric_capacity_mw,
        annual_energy_cost, annual_co2_tonnes, annual_carbon_cost,
        annual_om_cost, total_capex, annual_savings, payback, npv,
        GAS, ELECTRIC, CARBON, CAPEX, FINANCE, PLANT
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Steel TEA Summary"

    # Styles
    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def la(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def ra(): return Alignment(horizontal="right", vertical="center")
    def ca(): return Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16

    # Title
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "HyperHeat TEA — Steel Reheating Furnace (2025)"
    t.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.fill = fill("1F4E79"); t.alignment = ca()

    rows = [
        ("─── INPUTS ───", "", "", ""),
        ("Steel Output", f"{PLANT['annual_production_tonnes']:,}", "tonnes/year", "World Steel Assoc."),
        ("Heat Demand/Tonne", f"{PLANT['heat_demand_gj_per_tonne']}", "GJ/tonne", "IEA 2023"),
        ("Gas Efficiency", f"{GAS['efficiency']:.0%}", "", "IEA / Eurofer"),
        ("Electric Efficiency", f"{ELECTRIC['efficiency']:.0%}", "", "HyperHeat spec"),
        ("Gas Price", f"€{GAS['price_eur_per_kwh']}", "EUR/kWh", "Eurostat TTF H1 2025"),
        ("Electricity Price", f"€{ELECTRIC['price_eur_per_kwh']}", "EUR/kWh", "Eurostat H1 2025"),
        ("EU ETS Price", f"€{CARBON['ets_price_2025']}", "EUR/t CO2", "ICE EU ETS 2025"),
        ("─── CURRENT CASE ───", "", "", ""),
        ("Annual Gas Energy Bill", f"€{annual_energy_cost(GAS)/1e6:,.2f}M", "EUR/year", ""),
        ("Annual Electricity Bill", f"€{annual_energy_cost(ELECTRIC)/1e6:,.2f}M", "EUR/year", ""),
        ("Gas Carbon Fine", f"€{annual_carbon_cost(GAS, CARBON['ets_price_2025'])/1e6:,.2f}M", "EUR/year", ""),
        ("Electric Carbon Fine", f"€{annual_carbon_cost(ELECTRIC, CARBON['ets_price_2025'])/1e6:,.2f}M", "EUR/year", ""),
        ("Annual Saving (Current)", f"€{annual_savings(CARBON['ets_price_2025'])/1e6:,.2f}M", "EUR/year", "Negative = electric costs more"),
        ("Total CAPEX", f"€{total_capex()/1e6:,.2f}M", "EUR", "One-time investment"),
        ("Payback (Current)", str(payback(CARBON['ets_price_2025'])) if payback(CARBON['ets_price_2025']) else "❌ Not viable", "years", ""),
        ("NPV (Current)", f"€{npv(CARBON['ets_price_2025'])/1e6:,.1f}M", "EUR", "❌ Not viable at grid prices"),
        ("─── BEST CASE (PPA + ETS 2030) ───", "", "", ""),
    ]

    # Best case
    ELECTRIC["price_eur_per_kwh"] = 0.07
    savings_best = annual_savings(CARBON["ets_price_2030"])
    pb_best = payback(CARBON["ets_price_2030"])
    npv_best = npv(CARBON["ets_price_2030"])

    rows += [
        ("Electricity Price (PPA)", "€0.07", "EUR/kWh", "Renewable PPA deal"),
        ("ETS Price (2030 forecast)", f"€{CARBON['ets_price_2030']}", "EUR/t CO2", "BNEF / GMK Center consensus"),
        ("Annual Saving (Best Case)", f"€{savings_best/1e6:,.2f}M", "EUR/year", ""),
        ("Payback (Best Case)", f"{pb_best:.1f}", "years", "✅ Under 7 years = good investment"),
        ("NPV (Best Case)", f"€{npv_best/1e6:,.1f}M", "EUR", "✅ Strong positive NPV"),
    ]

    for i, (label, val, unit, note) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 20
        is_section = label.startswith("───")
        is_good = "✅" in str(val) or "✅" in str(note)

        for ci, v, bg in [
            (1, label, "2E75B6" if is_section else ("E2EFDA" if is_good else "F2F2F2")),
            (2, val, "2E75B6" if is_section else ("E2EFDA" if is_good else "FFFFFF")),
            (3, unit, "2E75B6" if is_section else "FFFFFF"),
            (4, note, "2E75B6" if is_section else "D9E1F2"),
        ]:
            c = ws.cell(row=i, column=ci, value=v)
            c.font = Font(name="Arial", bold=is_section, color="FFFFFF" if is_section else "000000", size=10)
            c.fill = fill(bg); c.border = bd()
            c.alignment = la() if ci in (1, 4) else ra()

    output_path = os.path.join(os.path.dirname(__file__), '..', 'excel', 'Steel_TEA_2025.xlsx')
    wb.save(output_path)
    print(f"  ✅ Saved: excel/Steel_TEA_2025.xlsx")


def build_cement():
    print("\n  Building Cement TEA Excel...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import sys, os

    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'models'))
    from cement_tea import (
        annual_energy_cost, fuel_co2_tonnes, annual_carbon_cost_fuel_only,
        total_capex, annual_savings, payback, npv, fuel_co2_reduction,
        pct_of_total_co2_saved, clinker_per_year, electric_capacity_mw,
        GAS, ELECTRIC, CARBON, CAPEX, PLANT, CO2_PER_TONNE_CEMENT
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cement TEA Summary"

    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def la(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def ra(): return Alignment(horizontal="right", vertical="center")
    def ca(): return Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 36

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "HyperHeat TEA — Cement Rotary Kiln (2025)"
    t.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.fill = fill("1F4E79"); t.alignment = ca()

    # Warning row
    ws.row_dimensions[2].height = 40
    ws.merge_cells("A2:D2")
    w = ws["A2"]
    w.value = "⚠️  KEY INSIGHT: 60% of cement CO₂ is from CALCINATION (chemistry). Cannot be avoided without CCS. HyperHeat eliminates the remaining 40% from fuel combustion."
    w.font = Font(name="Arial", italic=True, size=10, color="000000")
    w.fill = fill("FCE4D6"); w.alignment = la()

    rows = [
        ("─── INPUTS ───", "", "", ""),
        ("Cement Output", f"{PLANT['cement_output_tonnes']:,}", "tonnes/year", "Large EU plant benchmark"),
        ("Clinker Ratio", f"{PLANT['clinker_ratio']:.0%}", "", "Cembureau 2024 EU avg"),
        ("Heat Demand/Tonne Clinker", f"{PLANT['heat_gj_per_tonne_clinker']}", "GJ/tonne", "IEA Cement 2023"),
        ("Gas Kiln Efficiency", f"{GAS['efficiency']:.0%}", "", "IEA / Cembureau"),
        ("Electric Kiln Efficiency", f"{ELECTRIC['efficiency']:.0%}", "", "HyperHeat spec (conservative)"),
        ("Gas Price", f"€{GAS['price_eur_per_kwh']}", "EUR/kWh", "Eurostat TTF H1 2025"),
        ("Electricity Price", f"€{ELECTRIC['price_eur_per_kwh']}", "EUR/kWh", "Eurostat H1 2025"),
        ("─── CO₂ SPLIT (unique to cement) ───", "", "", ""),
        ("Calcination CO₂/tonne cement", f"{CO2_PER_TONNE_CEMENT['calcination']}", "t CO₂/t cement", "UNAVOIDABLE — IPCC AR6"),
        ("Fuel CO₂/tonne cement", f"{CO2_PER_TONNE_CEMENT['fuel']}", "t CO₂/t cement", "WHAT ELECTRIC ELIMINATES"),
        ("─── CURRENT CASE ───", "", "", ""),
        ("Annual Saving (Current)", f"€{annual_savings(CARBON['ets_price_2025'])/1e6:,.2f}M", "EUR/year", "Negative = electric costs more today"),
        ("Total CAPEX", f"€{total_capex()/1e6:,.2f}M", "EUR", "One-time investment"),
        ("Payback (Current)", str(payback(CARBON['ets_price_2025'])) if payback(CARBON['ets_price_2025']) else "❌ Not viable", "years", ""),
        ("NPV (Current)", f"€{npv(CARBON['ets_price_2025'])/1e6:,.1f}M", "EUR", "❌ Not viable at grid prices"),
        ("─── BEST CASE (PPA + ETS 2030) ───", "", "", ""),
    ]

    ELECTRIC["price_eur_per_kwh"] = 0.07
    savings_best = annual_savings(CARBON["ets_price_2030"])
    pb_best = payback(CARBON["ets_price_2030"])
    npv_best = npv(CARBON["ets_price_2030"])

    rows += [
        ("Electricity Price (PPA)", "€0.07", "EUR/kWh", "Renewable PPA deal"),
        ("ETS 2030 Forecast", f"€{CARBON['ets_price_2030']}", "EUR/t CO2", "BNEF / GMK Center consensus"),
        ("Annual Saving (Best Case)", f"€{savings_best/1e6:,.2f}M", "EUR/year", ""),
        ("Payback (Best Case)", f"{pb_best:.1f}" if pb_best else "❌", "years", "✅ Under 7 years = good" if pb_best and pb_best < 7 else ""),
        ("NPV (Best Case)", f"€{npv_best/1e6:,.1f}M", "EUR", "✅ Positive NPV" if npv_best > 0 else ""),
        ("─── CO₂ IMPACT ───", "", "", ""),
        ("Annual Fuel CO₂ Saved", f"{fuel_co2_reduction():,.0f}", "tonnes/year", ""),
        ("As % of Total Plant CO₂", f"{pct_of_total_co2_saved():.1f}%", "", "Rest requires CCS"),
        ("Lifetime CO₂ Saved", f"{fuel_co2_reduction()*15:,.0f}", "tonnes", "Over 15 year project life"),
    ]

    for i, (label, val, unit, note) in enumerate(rows, start=3):
        ws.row_dimensions[i].height = 22
        is_section = label.startswith("───")
        is_good = "✅" in str(val) or "✅" in str(note)

        for ci, v, bg in [
            (1, label, "2E75B6" if is_section else ("E2EFDA" if is_good else "F2F2F2")),
            (2, val, "2E75B6" if is_section else ("E2EFDA" if is_good else "FFFFFF")),
            (3, unit, "2E75B6" if is_section else "FFFFFF"),
            (4, note, "2E75B6" if is_section else "D9E1F2"),
        ]:
            c = ws.cell(row=i, column=ci, value=v)
            c.font = Font(name="Arial", bold=is_section, color="FFFFFF" if is_section else "000000", size=10)
            c.fill = fill(bg); c.border = bd()
            c.alignment = la() if ci in (1, 4) else ra()

    output_path = os.path.join(os.path.dirname(__file__), '..', 'excel', 'Cement_TEA_2025.xlsx')
    wb.save(output_path)
    print(f"  ✅ Saved: excel/Cement_TEA_2025.xlsx")


if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("  Building HyperHeat TEA Excel Models")
    print("=" * 50)
    check_openpyxl()
    build_steel()
    build_cement()
    print("\n  ✅ All done!\n")
